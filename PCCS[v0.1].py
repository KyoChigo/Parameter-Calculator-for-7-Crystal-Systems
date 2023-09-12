import openpyxl
import numpy as np

print("Parameter Calculator for 7 Crystal Systems [PCCS]")
print("    Author: Xu Zhihao (https://kyochigo.com)")
print("    Version: 0.1")

print("Please choose input mode: ")
print("    1: input by excel file (.xlsx)")
print("    2: input by keyboard")
mode = str(input())

if mode == "1":
    try:
        excel = openpyxl.load_workbook('template.xlsx')
        book = excel.worksheets[0]
        cryst_system = book.cell(3, 2).value
        if cryst_system in ["c", "t", "o", "h"]:
            m = np.array([])
            A = np.array([])
            for i in range(20):
                if book.cell(6+i, 1).value == None:
                    break
                m = np.append(m, book.cell(6+i, 1).value)
                A = np.append(A, np.array([book.cell(6+i, 2).value,
                                        book.cell(6+i, 3).value,
                                        book.cell(6+i, 4).value])**2)
            m = m**(-2)
            A = np.matrix(A.reshape(len(A)//3, 3))

        #elif cryst_system == "m":
            #to be updated in v0.2
        #elif cryst_system == "a":
            #to be updated in v0.2

    except FileNotFoundError:
        print("Please confirm that the template excel file has been fully uploaded to this page.")

elif mode == "2":
    print("Please choose crystal system: ")
    print("    c: cubic")
    print("    t: tetragonal")
    print("    o: orthorhombic")
    print("    h: hexagonal")
    print("    m: monoclinic")
    print("    a: triclinic")
    cryst_system = str(input())

    if cryst_system in ["c", "t", "o", "h"]:
        m = np.array([])
        A = np.array([])
        print("Please input values of d-spacing, h, k, l:")
        print("   Input 0 when you finish.")
        while True:
            temp = list(float(x) for x in input().split())
            if temp == [0]:
                m = m**(-2)
                A = A.reshape(len(A)//3, 3)
                break
            elif len(temp) == 4:
                m = np.append(m, temp[0])
                A = np.append(A, temp[1:4])
            else:
                print("Invalid value!")

    #elif cryst_system == "m":
        #to be updated in v0.2
    #elif cryst_system == "a":
        #to be updated in v0.2

print("Calculation result:")
if cryst_system == "c":
    X = np.dot(np.matmul(np.matmul(A.T, A)**(-1), A.T), m)
    a = ((1/X[0, 0])**0.5 + (1/X[0, 1])**0.5 + (1/X[0, 2])**0.5)/3
    print("a: " + '{:.3f}'.format(a))

elif cryst_system == "t":
    X = np.dot(np.matmul(np.matmul(A.T, A)**(-1), A.T), m)
    a = (1/X[0, 0])**0.5
    b = (1/X[0, 1])**0.5
    print("a: " + '{:.3f}'.format(a))
    print("b: " + '{:.3f}'.format(b))

elif cryst_system == "o":
    X = np.dot(np.matmul(np.matmul(A.T, A)**(-1), A.T), m)
    a = (1/X[0, 0])**0.5
    b = (1/X[0, 1])**0.5
    c = (1/X[0, 2])**0.5
    print("a: " + '{:.3f}'.format(a))
    print("b: " + '{:.3f}'.format(b))
    print("c: " + '{:.3f}'.format(c))
